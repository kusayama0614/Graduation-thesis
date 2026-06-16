import json
import csv
import os
from typing import Tuple, Dict, Any, List

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


def _normalize_key_dict(d: Dict[Any, Any]) -> Dict[str, str]:
    out = {}
    for k, v in d.items():
        out[str(k).strip()] = str(v).strip()
    return out


def parse_csv(filepath: str) -> Dict[str, Any]:
    candidates = []
    with open(filepath, newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        rows = [r for r in reader if any(cell.strip() for cell in r)]

    # Try header detection: header like question,answer
    if rows:
        first = [c.strip().lower() for c in rows[0]]
        if 'question' in first and 'answer' in first:
            qidx = first.index('question')
            aidx = first.index('answer')
            for r in rows[1:]:
                if len(r) > max(qidx, aidx):
                    candidates.append({'question': r[qidx].strip(), 'answer': r[aidx].strip()})
            return {'success': True, 'candidates': candidates, 'total_candidates': len(candidates)}

    # Fallback: assume two columns question,answer
    for r in rows:
        if len(r) >= 2:
            candidates.append({'question': r[0].strip(), 'answer': r[1].strip()})

    return {'success': True, 'candidates': candidates, 'total_candidates': len(candidates)}


def parse_xlsx(filepath: str) -> Dict[str, Any]:
    if load_workbook is None:
        return {'success': False, 'error': 'openpyxl not installed', 'candidates': [], 'total_candidates': 0}
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c).strip() if c is not None else '' for c in row])

    # reuse csv parser logic by writing to temp csv-like structure
    candidates = []
    if rows:
        first = [c.strip().lower() for c in rows[0]]
        if 'question' in first and 'answer' in first:
            qidx = first.index('question')
            aidx = first.index('answer')
            for r in rows[1:]:
                if len(r) > max(qidx, aidx):
                    candidates.append({'question': r[qidx].strip(), 'answer': r[aidx].strip()})
            return {'success': True, 'candidates': candidates, 'total_candidates': len(candidates)}

    for r in rows:
        if len(r) >= 2:
            candidates.append({'question': r[0].strip(), 'answer': r[1].strip()})

    return {'success': True, 'candidates': candidates, 'total_candidates': len(candidates)}


def parse_json(filepath: str) -> Dict[str, Any]:
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Accept formats: {"answers": {"1":"A"}} or [{"question":1,"answer":"A"}, ...]
    if isinstance(data, dict):
        if 'answers' in data and isinstance(data['answers'], dict):
            candidates = [{'question': k, 'answer': v} for k, v in data['answers'].items()]
            return {'success': True, 'candidates': candidates, 'total_candidates': len(candidates)}
        # If dict of question->answer directly
        if all(isinstance(v, (str, int, float)) for v in data.values()):
            candidates = [{'question': k, 'answer': v} for k, v in data.items()]
            return {'success': True, 'candidates': candidates, 'total_candidates': len(candidates)}

    if isinstance(data, list):
        candidates = []
        for item in data:
            if isinstance(item, dict) and 'question' in item and 'answer' in item:
                candidates.append({'question': item['question'], 'answer': item['answer']})
        return {'success': True, 'candidates': candidates, 'total_candidates': len(candidates)}

    return {'success': False, 'error': 'Unrecognized JSON format', 'candidates': [], 'total_candidates': 0}


def parse_answer_key_file(filepath: str, filename: str) -> Dict[str, Any]:
    ext = os.path.splitext(filename)[1].lower()
    if ext in ('.csv',):
        parsed = parse_csv(filepath)
        parsed.update(validate_candidates(parsed.get('candidates', [])))
        return parsed
    if ext in ('.xlsx', '.xlsm', '.xltx'):
        parsed = parse_xlsx(filepath)
        parsed.update(validate_candidates(parsed.get('candidates', [])))
        return parsed
    if ext in ('.json',):
        parsed = parse_json(filepath)
        parsed.update(validate_candidates(parsed.get('candidates', [])))
        return parsed

    return {'success': False, 'error': 'Unsupported file type', 'candidates': [], 'total_candidates': 0, 'errors': [], 'warnings': []}


def validate_candidates(candidates: List[Dict[str, Any]]) -> Dict[str, Any]:
    """基本的な検証: 空エントリ、重複、回答が空の項目を検出して errors/warnings を返す"""
    errors: List[str] = []
    warnings: List[str] = []
    seen = {}
    for idx, item in enumerate(candidates):
        q = str(item.get('question', '')).strip()
        a = str(item.get('answer', '')).strip()
        if not q:
            errors.append(f'行 {idx+1}: 問題番号が空です')
        if not a:
            warnings.append(f'行 {idx+1}: 解答が空です')
        if q:
            if q in seen:
                errors.append(f'重複した問題番号: {q} (行 {seen[q]} と {idx+1})')
            else:
                seen[q] = idx+1

    # If numeric question keys, check for gaps starting from 1
    numeric_keys = []
    for k in seen.keys():
        try:
            numeric_keys.append(int(str(k)))
        except Exception:
            numeric_keys = []
            break

    if numeric_keys:
        numeric_keys_sorted = sorted(numeric_keys)
        expected = list(range(1, numeric_keys_sorted[-1] + 1))
        missing = [str(n) for n in expected if n not in numeric_keys_sorted]
        if missing:
            warnings.append(f'欠番があります: {",".join(missing)}')

    return {'errors': errors, 'warnings': warnings}
